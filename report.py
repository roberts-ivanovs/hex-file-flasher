import click
import sqlite3
from openpyxl import Workbook
from openpyxl import styles

@click.command()
@click.option(
    "--start-date", required=True, type=click.DateTime(formats=["%Y-%m-%d_%H:%M"])
)
@click.option(
    "--end-date", required=True, type=click.DateTime(formats=["%Y-%m-%d_%H:%M"])
)
def generate_report(start_date, end_date):

    conn = sqlite3.connect("db.sqlite3")
    c = conn.cursor()
    data = c.execute(
        "SELECT chip_type, software, chip_number, MAX(flash.id), flash.flashed_id FROM flash "
        "JOIN chip ON flash.chip_fk=chip.id "
        "WHERE flash.flashed_time>? AND flash.flashed_time<? GROUP BY chip.chip_number "
        "ORDER BY flash.flashed_id",
        [start_date.timestamp(), end_date.timestamp()],
    ).fetchall()

    # Create the actual data sheet
    wb = Workbook()

    # Remove the sheet created by default
    wb.remove(wb["Sheet"])

    # Set a fill colour to orange
    fill = styles.PatternFill("solid", fgColor="FD5421")

    titlebars = {
        "Chip type": "A",
        "Code type": "B",
        "Chip id": "C",
        "Rssi": "D",
        "Flashed": "E",
        "Flashed ID": "F",
        "Notes": "G",
        None: "H",
        "dB vs best": "I",
        "Pass": "J",
    }

    # Create a worksheet and select it
    ws = wb.create_sheet("REPORT")
    # Write out the title bar
    ws.append([i for i in titlebars.keys()])
    dev_ln = len(data)

    # Bottom-most formulas/calculations
    # Set the equations for the xlsx file
    eq1 = f"""=AVERAGEIF({titlebars["Rssi"]}2:{titlebars["Rssi"]}{dev_ln + 1},CONCATENATE(">", MAX({titlebars["Rssi"]}2:{titlebars["Rssi"]}{dev_ln + 1})-5))"""
    eq2 = f"""=COUNTIF({titlebars['Pass']}$2:{titlebars['Pass']}{dev_ln + 1}, "PASS")"""
    eq3 = f"={titlebars['dB vs best']}{ dev_ln + 5}/(ROW({titlebars['Pass']}{dev_ln + 1})-1)"

    # Get the according coordinates
    txt_top_5db = titlebars["Notes"] + str(dev_ln + 3)
    txt_failed_cnt = titlebars["Notes"] + str(dev_ln + 5)
    txt_failed_rate = titlebars["Notes"] + str(dev_ln + 6)
    calc_top_5db = titlebars['dB vs best'] + str(dev_ln + 3)
    calc_failed_cnt = titlebars['dB vs best'] + str(dev_ln + 5)
    calc_failed_rate = titlebars['dB vs best'] + str(dev_ln + 6)

    # Set the text for each coordinate
    ws[txt_top_5db] = "Top 5dB average:"
    ws[txt_failed_cnt] = "Succeeded units:"
    ws[txt_failed_rate] = "Succeeded rate:"

    ws[txt_top_5db].alignment = styles.Alignment(horizontal='right')
    ws[txt_failed_cnt].alignment = styles.Alignment(horizontal='right')
    ws[txt_failed_rate].alignment = styles.Alignment(horizontal='right')

    # Insert the previously defined equations
    ws[calc_top_5db] = eq1
    ws[calc_failed_cnt] = eq2
    ws[calc_failed_rate] = eq3

    # Extra visual formating
    ws.freeze_panes = 'A2'
    ws.column_dimensions[titlebars['Notes']].width = 40
    ws.column_dimensions[titlebars['dB vs best']].width = 10
    ws.column_dimensions[titlebars[None]].width = 1


    # Insert the DB data
    for row, entry in enumerate(data):
        row = str(row + 2)
        cell = {i[0]: i[1] + row for i in titlebars.items()}

        # Get test results
        test_data = c.execute(
            "SELECT key, value from test WHERE flash_fk=?", [entry[3]]
        ).fetchall()

        # Insert test results
        got_flashed = False
        got_rssi = False
        for test_data_entry in test_data:
            if test_data_entry[0] == "flashed":
                ws[cell["Flashed"]] = test_data_entry[1]
                if test_data_entry[1] == "false":
                    ws[cell["Flashed"]].fill = fill
                got_flashed = True
            elif test_data_entry[0] == "rssi":
                ws[cell["Rssi"]] = int(test_data_entry[1])
                got_rssi = True

        if not got_rssi:
            ws[cell["Rssi"]] = "n/a"
            ws[cell["Rssi"]].fill = fill

        if not got_flashed:
            ws[cell["Flashed"]] = "False"
            ws[cell["Flashed"]].fill = fill

        # Insert base results
        ws[cell["Chip type"]] = entry[0]
        ws[cell["Code type"]] = entry[1]
        ws[cell["Chip id"]] = entry[2]
        ws[cell["Flashed ID"]] = entry[4]

        rounddown = f"=ROUNDDOWN({cell['Rssi']}-${calc_top_5db}, -1)"
        calc_pass = f"=IF(AND({cell['Flashed']}=\"true\", {cell['dB vs best']}>=-10), \"PASS\", \"NO PASS\")"
        ws[cell["dB vs best"]] = rounddown
        ws[cell["Pass"]] = calc_pass

    name = f"report {start_date}-{end_date}.xlsx"
    wb.save(name)


if __name__ == "__main__":
    generate_report()
